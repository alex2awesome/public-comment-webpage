#!/usr/bin/env python3
"""
Step 1: Add columns and rows to FOIA tracker spreadsheet.

Updates:
- 3 new columns: AI Records Retention Policy, Prior FOIA Precedent, Template Adaptation Notes
- 9 new rows: 8 states + Kent, WA
- Fill in new columns for all existing rows

Reads foia_tracker_detailed_v2.xlsx, writes foia_tracker_detailed_v3.xlsx.
"""

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from copy import copy

from helpers import (
    load_tracker,
    save_tracker,
    build_jur_map,
    update_cell,
    COL_RETENTION,
    COL_PRECEDENT,
    COL_TEMPLATE_NOTES,
)

wb, ws = load_tracker("foia_tracker_detailed_v2.xlsx")

# =============================================================================
# STEP 1: Add 3 new column headers
# =============================================================================
NEW_COL_RETENTION = COL_RETENTION    # 23 / W
NEW_COL_PRECEDENT = COL_PRECEDENT    # 24 / X
NEW_COL_TEMPLATE = COL_TEMPLATE_NOTES  # 25 / Y

ws.cell(row=1, column=NEW_COL_RETENTION, value="AI Records Retention Policy")
ws.cell(row=1, column=NEW_COL_PRECEDENT, value="Prior FOIA Precedent")
ws.cell(row=1, column=NEW_COL_TEMPLATE, value="Template Adaptation Notes")

# Copy header formatting from adjacent column
header_font = Font(bold=True)
for col in [NEW_COL_RETENTION, NEW_COL_PRECEDENT, NEW_COL_TEMPLATE]:
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.alignment = Alignment(wrap_text=True)

# =============================================================================
# STEP 2: Insert 8 new state rows after row 25 (New Hampshire), before cities
# =============================================================================
# Insert 8 rows at position 26, pushing cities down
INSERT_STATE_ROW = 26
NUM_STATE_ROWS = 8
ws.insert_rows(INSERT_STATE_ROW, NUM_STATE_ROWS)

# New state data (rows 26-33 after insertion)
new_states = [
    {  # Row 26: Washington State (Tier 1)
        "Cat": "US-State",
        "Tier": 1,
        "Jurisdiction": "Washington",
        "Country": "US",
        "Level": "State",
        "Known AI Deployment": (
            "EMPLOYEE-FACING: Governor Inslee signed EO 24-01 (Jan 30, 2024) directing WaTech to develop GenAI guidelines. "
            "WaTech published interim GenAI guidelines (Aug 2023) and 'State of Washington Generative AI Report' (Oct 2024). "
            "Microsoft Copilot for Government being adopted at agency level. WaTech EA-01-02-S policy governs responsible GenAI use. "
            "State agencies using GenAI for drafting docs, summarizing text, plain-language rewrites, auto-translation. "
            "OFM assessed GenAI workforce impact (Dec 2024). WA AI Task Force issued interim report (Dec 2025). "
            "| Sources: https://watech.wa.gov/artificial-intelligence-resources | "
            "https://governor.wa.gov/sites/default/files/exe_order/24-01%20-%20Artificial%20Intelligence%20(tmp).pdf | "
            "https://watech.wa.gov/sites/default/files/2024-10/WA_State_GenAIReport_FINAL.pdf"
        ),
        "Est. Employees": "~77,000",
        "Proxy?": "No",
        "Filing Method": (
            "Each state agency has its own public records officer. No centralized portal. "
            "Requests can be made in person, by mail, or email. No specific form required.\n"
            "WATECH: Email PublicRecords@watech.wa.gov; Mail: Attn: Public Records/Records Mgmt. Unit, "
            "WaTech, PO Box 41454, Olympia, WA 98501; In person: 1500 Jefferson St SE, Olympia, WA 98501\n"
            "Many agencies use NextRequest portals."
        ),
        "Source URL": "https://watech.wa.gov/about/public-records-requests",
        "Cost": "$0.15/page photocopies; $0.10/page scanning; $0.05 per 4 electronic files; flat fee up to $2.00/request (RCW 42.56.120)",
        "Statute": "Washington Public Records Act, RCW 42.56",
        "Response Deadline": "5 business days to respond (not counting day of receipt); must acknowledge and provide estimate (RCW 42.56.520)",
        "Expected Turnaround": "5-30 days",
        "Template #": 1,
        "GovAI?": "-",
        "Notes": "HIGHEST PRIORITY STATE GAP. 6 WA cities already in tracker but no state entry. WA PRA is one of the strongest in the US — no exemptions for 'burden.'",
    },
    {  # Row 27: Missouri (Tier 2)
        "Cat": "US-State",
        "Tier": 2,
        "Jurisdiction": "Missouri",
        "Country": "US",
        "Level": "State",
        "Known AI Deployment": (
            "EMPLOYEE-FACING: Governor Kehoe signed EO 26-02 (Jan 2026) directing strategic AI framework. "
            "Office of Administration oversees AI for 14 state agencies. Tim Marczewski is Director of AI and Innovation. "
            "Uses Google Gemini in government-secured tenant. Three internal chatbot prototypes in development. "
            "'Ask Mo' external chatbot for citizen services. All AI with human-in-the-loop policy. "
            "| Sources: https://www.newstribune.com/news/2025/dec/14/missouri-officials-using-ai-to-enhance-operations/ | "
            "https://www.missourinet.com/2025/11/12/missouri-office-of-administrations-ai-tools-are-cyber-secure-says-state-ai-leader/ | "
            "https://www.govtech.com/artificial-intelligence/missouri-executive-order-seeks-to-advance-ai-strategy-use"
        ),
        "Est. Employees": "~103,500",
        "Proxy?": "No",
        "Filing Method": (
            "No centralized state portal. Requests go to custodian at each agency.\n"
            "AG's Office: sunshinerequest@ago.mo.gov\n"
            "DOR online form: https://dor.mo.gov/sunshine-request/request.html\n"
            "AG sample request form: https://ago.mo.gov/get-help/programs-services-from-a-z/sunshine-law/sample-language-forms/records-request-form/"
        ),
        "Source URL": "https://ago.mo.gov/get-help/programs-services-from-a-z/sunshine-law/",
        "Cost": "~$0.10/page copying; reasonable research/production fees vary by agency",
        "Statute": "Missouri Sunshine Law, Chapter 610 RSMo (RSMo 610.010 et seq.)",
        "Response Deadline": "3 business days (RSMo 610.023.3); must provide written explanation if delayed",
        "Expected Turnaround": "3-10 biz days",
        "Template #": 1,
        "GovAI?": "-",
        "Notes": "3-day deadline is one of the tightest in the country. Penalties: up to $5,000 for purposeful violations (RSMo 610.027) + attorney fees.",
    },
    {  # Row 28: Virginia (Tier 2)
        "Cat": "US-State",
        "Tier": 2,
        "Jurisdiction": "Virginia",
        "Country": "US",
        "Level": "State",
        "Known AI Deployment": (
            "EMPLOYEE-FACING: Governor Youngkin signed EO 30 (2024) establishing AI baseline standards and AI Task Force. "
            "VITA published AI policy standards (June 2024). Microsoft Copilot Chat available to COV employees; "
            "VITA hosts regular Copilot training workshops. First state-level agentic AI tool for regulatory review (EO 51, July 2025). "
            "ODGA / Chief Data Officer coordinates AI strategy. Governor vetoed comprehensive AI regulation (HB 2094, 2025). "
            "| Sources: https://www.vita.virginia.gov/news--events/latest-events/name-1057122-en.html | "
            "https://www.governor.virginia.gov/newsroom/news-releases/2025/july/name-1053152-en.html | "
            "https://www.govtech.com/artificial-intelligence/with-plans-for-all-sectors-virginia-aims-to-get-ahead-of-ai"
        ),
        "Est. Employees": "~162,700",
        "Proxy?": "No",
        "Filing Method": (
            "Requests go to FOIA officer at each public body. No centralized portal.\n"
            "VITA: https://www.vita.virginia.gov/about/foia/ ($45/hr time charges)\n"
            "Governor's Office: FOIA@governor.virginia.gov ('FOIA Request' in subject)\n"
            "Each public body must designate a FOIA officer with public contact info."
        ),
        "Source URL": "https://foiacouncil.dls.virginia.gov/",
        "Cost": "Reasonable charges not exceeding actual cost; must itemize if >$200; VITA charges $45/hr",
        "Statute": "Virginia FOIA, Code of Virginia § 2.2-3700 et seq.",
        "Response Deadline": "5 working days; may extend 7 more with written notice (§ 2.2-3704)",
        "Expected Turnaround": "5-12 working days",
        "Template #": 1,
        "GovAI?": "-",
        "Notes": "Civil penalties: $500-$2,000 first offense; $2,000-$5,000 subsequent (§ 2.2-3714). Active AI governance with Chief Data Officer.",
    },
    {  # Row 29: Indiana (Tier 2)
        "Cat": "US-State",
        "Tier": 2,
        "Jurisdiction": "Indiana",
        "Country": "US",
        "Level": "State",
        "Known AI Deployment": (
            "EMPLOYEE-FACING: AI Policy v1.1 (Dec 2024) via OCDO. ChatGPT broadly permitted for state employees "
            "(exempt from full review for web-based GenAI). 'Captain Record' (Google Gemini 2.0/Vertex AI): "
            "SoS GenAI tool searching 20M+ pages of archived records. 'Ask Indiana' AI assistant on in.gov. "
            "DWD Uplink AI career recommendation tool (production Nov 2023). IOT security log AI (production 2021). "
            "SB 150 (March 2024) created AI Task Force. NIST AI RMF adopted. "
            "| Sources: https://www.in.gov/mph/AI/ | https://statescoop.com/indiana-captain-record-ai-search-tool/ | "
            "https://www.govtech.com/biz/data/indianas-new-ai-policy-calls-for-pre-deployment-assessments"
        ),
        "Est. Employees": "~32,000",
        "Proxy?": "No",
        "Filing Method": (
            "Requests go to specific agency. Can be made by phone, mail, in person.\n"
            "Standard form: https://forms.in.gov/download.aspx?id=11622\n"
            "INDOT: https://www.in.gov/indot/public-involvement/indot-public-records-request-form/\n"
            "Public Access Counselor: in.gov/pac (enforcement/complaints)"
        ),
        "Source URL": "https://www.in.gov/pac/",
        "Cost": "Free inspection; email copies may be free; max $0.10/page paper; NO fees for labor, search, or review (IC 5-14-3-8)",
        "Statute": "Indiana Access to Public Records Act (APRA), IC 5-14-3",
        "Response Deadline": "24 hrs (in-person/phone); 7 calendar days (mail/fax/email) for acknowledgment; 'reasonable time' for production",
        "Expected Turnaround": "1-2 weeks (simple); 2-6 weeks (complex)",
        "Template #": 1,
        "GovAI?": "-",
        "Notes": "No fees for electronic records is very favorable. Active Public Access Counselor for enforcement.",
    },
    {  # Row 30: Montana (Tier 2)
        "Cat": "US-State",
        "Tier": 2,
        "Jurisdiction": "Montana",
        "Country": "US",
        "Level": "State",
        "Known AI Deployment": (
            "EMPLOYEE-FACING: Early stage. Commissioner of Political Practices used ChatGPT to draft official opinion on robocalls. "
            "MDT using neural network algorithms for contract time estimation. EO 5-2025 (Aug 2025) directs DLI to expand AI training. "
            "Hiring first Chief AI Officer. HB 178 (signed May 2025) limits government AI use, requires disclosure. "
            "SB 212 (April 2025) 'Right to Compute Act' — first-in-nation. 48 AI-related bills introduced. "
            "| Sources: https://www.govtech.com/artificial-intelligence/executive-order-targets-ai-readiness-for-montana-workforce | "
            "https://montanafreepress.org/2025/09/01/artificial-intelligence-offering-political-practices-advice-about-robocalls-in-montana-gop-internal-spat/ | "
            "https://www.govtech.com/workforce/help-wanted-montana-seeks-its-first-chief-ai-officer"
        ),
        "Est. Employees": "~29,000",
        "Proxy?": "No",
        "Filing Method": (
            "CENTRALIZED PORTAL: Office of Public Information Requests (OPIR)\n"
            "ONLINE: https://mt-pir.arkcase.us/pir/portal/reading-room\n"
            "EMAIL: publicrecords@mt.gov\n"
            "MAIL: OPIR, Dept of Administration, 125 N. Roberts St, PO Box 200101, Helena, MT 59620-0101\n"
            "PHONE: (406) 444-2686"
        ),
        "Source URL": "https://opir.mt.gov/records-request",
        "Cost": "First hour free; $25/hr after; $0.10/page copies; no fee waivers",
        "Statute": "Montana Constitution Art. II, Sec. 9 (Right to Know) + MCA Title 2, Ch. 6 (§ 2-6-1002 et seq.)",
        "Response Deadline": "5 business days acknowledgment (exec branch); 'timely manner' for others; 90-day max extension",
        "Expected Turnaround": "1-3 weeks (simple); 3-8 weeks (complex)",
        "Template #": 1,
        "GovAI?": "-",
        "Notes": "CONSTITUTIONAL right of access (one of ~6 states). Strong judicial enforcement leverage. Commissioner's documented ChatGPT use is a strong specific target.",
    },
    {  # Row 31: Nebraska (Tier 2)
        "Cat": "US-State",
        "Tier": 2,
        "Jurisdiction": "Nebraska",
        "Country": "US",
        "Level": "State",
        "Known AI Deployment": (
            "EMPLOYEE-FACING: NITC adopted AI policy section 8-609 (Nov 8, 2024) establishing framework for government AI use. "
            "NITC Technical Panel discussed Gemini and Copilot. NDE launched AI Introduction course (Feb 2025) for govt employees. "
            "LR430 (March 2024) studying AI impact. LB 77 banning AI as sole basis for prior authorization denial. "
            "| Sources: https://nitc.nebraska.gov/docs/8-609.pdf | https://events.govtech.com/Nebraska-Data-and-AI-Summit"
        ),
        "Est. Employees": "~44,000",
        "Proxy?": "No",
        "Filing Method": (
            "Requests go to custodian of specific record or agency head.\n"
            "Governor's Office: https://governor.nebraska.gov/public-records-request\n"
            "State Patrol: https://statepatrol.nebraska.gov/services/public-records-requests\n"
            "No centralized statewide portal."
        ),
        "Source URL": "https://ago.nebraska.gov/public-records",
        "Cost": "Actual material cost only; first 8 hours of staff search/ID/redaction/copying FREE; no attorney review fees",
        "Statute": "Nebraska Public Records Statutes, Neb. Rev. Stat. §§ 84-712 through 84-712.09",
        "Response Deadline": "4 business days (written requests); extensions permitted with written justification",
        "Expected Turnaround": "4-10 biz days",
        "Template #": 1,
        "GovAI?": "-",
        "Notes": "Historically ranked #1 FOI law nationally by BGA/IRE. 8 free hours of staff time + no attorney fees = very cost-effective.",
    },
    {  # Row 32: Maryland (Tier 2)
        "Cat": "US-State",
        "Tier": 2,
        "Jurisdiction": "Maryland",
        "Country": "US",
        "Level": "State",
        "Known AI Deployment": (
            "EMPLOYEE-FACING: Governor Moore signed EO 01.01.2024.02 (Jan 8, 2024) — 'Catalyzing Responsible and Productive Use of AI.' "
            "AI Governance Act (SB818) effective July 2024. AI Subcabinet led by Senior Advisor Nishant Shah. "
            "Google Gemini extended to ~43,000 state employees; ~12,500 actively using GenAI. "
            "Individual agencies also experimenting with ChatGPT, Claude, Gemini COTS. "
            "2025 AI Enablement Strategy covers 12 'critical domains.' Dedicated AI portal: https://ai.maryland.gov/ "
            "| Sources: https://doit.maryland.gov/policies/ai/Pages/maryland-AI-enablement-strategy-and-roadmap.aspx | "
            "https://www.govtech.com/artificial-intelligence/maryland-expands-ai-integration-across-state-government | "
            "https://ai.maryland.gov/"
        ),
        "Est. Employees": "~126,600",
        "Proxy?": "No",
        "Filing Method": (
            "Requests go to specific agency. Many agencies have online PIA forms.\n"
            "Comptroller: pia@marylandtaxes.gov\n"
            "DPSCS: dpscs.pia@maryland.gov\n"
            "MDOT: https://www.mdot.maryland.gov/tso/pages/Index.aspx?PageId=69\n"
            "AG PIA FAQ: https://oag.maryland.gov/resources-info/Pages/public-information-act-faq.aspx"
        ),
        "Source URL": "https://oag.maryland.gov/resources-info/Pages/public-information-act-faq.aspx",
        "Cost": "First 2 hours search/prep FREE; $0.25/page paper; fee waivers available on request",
        "Statute": "Maryland Public Information Act (MPIA), General Provisions Article §§ 4-101 through 4-601",
        "Response Deadline": "30 days to grant/deny; must notify within 10 working days if delayed",
        "Expected Turnaround": "10-30 days",
        "Template #": 1,
        "GovAI?": "-",
        "Notes": "Baltimore (city) already in tracker. Governor's AI EO + 12,500 active GenAI users = strong target. Todd Feathers reportedly targeting MD.",
    },
    {  # Row 33: Tennessee (Tier 2)
        "Cat": "US-State",
        "Tier": 2,
        "Jurisdiction": "Tennessee",
        "Country": "US",
        "Level": "State",
        "Known AI Deployment": (
            "EMPLOYEE-FACING: 1,000 ChatGPT Enterprise licenses (spring 2025), 60-day pilot then enterprise-wide rollout. "
            "Priority: shared services, communications, HR, operations, exec support, research, IT. "
            "Active users saving 1.5-2.5 hrs/day. STS manages AI strategy with formal AI Review Committee. "
            "AI Advisory Council 2025 Action Plan (approved Nov 17, 2025). Enterprise AI Policy 200-POL-007 and GenAI Policy ISC 3.00. "
            "Nashville Metro Schools deployed AI in all high schools (2025). "
            "| Sources: https://statescoop.com/tennessee-chatgpt-enterprise-kristin-darby/ | "
            "https://www.tn.gov/finance/ai-council.html | "
            "https://www.tn.gov/finance/strategic-technology-solutions/artificial-intelligence.html"
        ),
        "Est. Employees": "~99,300",
        "Proxy?": "No",
        "Filing Method": (
            "Requests go to Public Record Request Coordinator (PRRC) at each entity.\n"
            "Standard form (accepted by all entities): https://comptroller.tn.gov/office-functions/open-records-counsel.html\n"
            "Comptroller: https://comptroller.tn.gov/about-us/public-records-requests.html\n"
            "HR: https://www.tn.gov/hr/about/public-records-request.html\n"
            "May be submitted in person, phone, fax, mail, email, or online portal."
        ),
        "Source URL": "https://comptroller.tn.gov/office-functions/open-records-counsel.html",
        "Cost": "$0.15/page B&W; $0.50/page color; agencies may charge labor for producing copies",
        "Statute": "Tennessee Public Records Act, TCA § 10-7-503 et seq. (Constitutional basis: TN Const. Art. I, § 19)",
        "Response Deadline": "7 business days to produce, deny in writing, or provide estimated timeframe",
        "Expected Turnaround": "7-21 biz days",
        "Template #": 1,
        "GovAI?": "-",
        "Notes": "Constitutional basis for open records. ChatGPT Enterprise with 1,000 licenses = documented records. Office of Open Records Counsel for disputes.",
    },
]

# Column mapping
COL_MAP = {
    "Cat": 1, "Tier": 2, "Jurisdiction": 3, "Country": 4, "Level": 5,
    "Known AI Deployment": 6, "Est. Employees": 7, "Proxy?": 8,
    "Filing Method": 9, "Source URL": 10, "Cost": 11, "Statute": 12,
    "Response Deadline": 13, "Expected Turnaround": 14, "Template #": 15,
    "GovAI?": 16, "Notes": 21,
}

for i, state in enumerate(new_states):
    row = INSERT_STATE_ROW + i
    for field, col in COL_MAP.items():
        if field in state:
            ws.cell(row=row, column=col, value=state[field])
    # Set wrap text for long fields
    for col in [6, 9, 21]:
        ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True)

print(f"Inserted {len(new_states)} new state rows at rows {INSERT_STATE_ROW}-{INSERT_STATE_ROW + len(new_states) - 1}")

# =============================================================================
# STEP 3: Insert Kent, WA row near other WA cities
# After inserting 8 state rows, WA cities shifted down by 8:
# Bellingham was 29 -> 37, Everett 30 -> 38, Seattle 31 -> 39
# Bellevue 39 -> 47, Spokane 40 -> 48, Tacoma 41 -> 49
# Insert Kent after Tacoma (row 49), so at row 50
# =============================================================================
KENT_ROW = 50  # After Tacoma (now row 49)
ws.insert_rows(KENT_ROW, 1)

kent_data = {
    "Cat": "US-City",
    "Tier": 2,
    "Jurisdiction": "Kent, WA",
    "Country": "US",
    "Level": "City",
    "Known AI Deployment": (
        "EMPLOYEE-FACING: AI use confirmed via completed public records request — Kent PD records obtained by Rose Terse "
        "(MuckRock, completed ~30 days). Specific tools not yet publicly documented but records exist proving GenAI use. "
        "| Sources: MuckRock request by Rose Terse"
    ),
    "Est. Employees": "~2,200",
    "Proxy?": "No",
    "Filing Method": (
        "ONLINE: kentwa.gov/PublicRecords\n"
        "EMAIL: publicrecords@kentwa.gov\n"
        "MAIL: City Clerk, City of Kent, 220 Fourth Ave S, Kent, WA 98032\n"
        "PHONE: (253) 856-5725"
    ),
    "Source URL": "https://www.kentwa.gov/government/city-clerk/public-records",
    "Cost": "$0.15/page per RCW 42.56.120; electronic copies at actual cost",
    "Statute": "Washington Public Records Act, RCW 42.56",
    "Response Deadline": "5 business days (RCW 42.56.520)",
    "Expected Turnaround": "5-30 days (Kent PD responded in ~30 days to prior request)",
    "Template #": 1,
    "GovAI?": "-",
    "Notes": "PROVEN PRECEDENT: Rose Terse (MuckRock) obtained AI chat records from Kent PD in ~30 days. Strong data point for WA city requests.",
}

for field, col in COL_MAP.items():
    if field in kent_data:
        ws.cell(row=KENT_ROW, column=col, value=kent_data[field])
for col in [6, 9, 21]:
    ws.cell(row=KENT_ROW, column=col).alignment = Alignment(wrap_text=True)

print(f"Inserted Kent, WA at row {KENT_ROW}")

# =============================================================================
# STEP 4: Fill in AI Records Retention Policy (Col 23) for ALL rows
# After insertions: total rows shifted. Let's recalculate.
# Original 102 rows + 8 state rows + 1 Kent row = 111 data rows (rows 2-111 are now data, row 1 is header)
# =============================================================================

# Build jurisdiction -> row mapping (after all insertions)
jur_map = build_jur_map(ws)

print(f"\nTotal rows after insertions: {ws.max_row}")
print(f"Jurisdictions mapped: {len(jur_map)}")

# Retention policy data for all jurisdictions
retention_data = {
    # === KNOWN POLICIES (verified) ===
    "Seattle, WA": (
        "YES — POL-209 (Nov 3, 2023): First US city GenAI policy. "
        "§6.1: AI records are public records subject to disclosure. "
        "§6.2: Vendors must support data export for records requests. "
        "§6.3: Employees must retain AI inputs/outputs per retention schedules. "
        "| Source: https://www.seattle.gov/documents/Departments/InformationTechnology/City-of-Seattle-Generative-AI-Policy.pdf"
    ),
    "Connecticut": (
        "YES — Policy AI-01, AI Responsible Use Framework (Feb 2024). "
        "Governs state employee use of AI tools. Requires documentation of AI use in decision-making. "
        "AI Engagement & Enablement Lab for safe experimentation. "
        "| Source: https://portal.ct.gov/das/CTEdTech/Artificial-Intelligence"
    ),
    "Illinois": (
        "YES — DoIT AI Policy (effective April 1, 2025). "
        "Governs state employee AI use. Start with low-risk/high-reward use cases. "
        "Open-source AI preference. Documentation requirements for AI-assisted decisions. "
        "| Source: https://doit.illinois.gov/ai"
    ),
    "North Carolina": (
        "YES — Copilot for M365 policy (updated Feb 26, 2025). "
        "Statewide pilot through June 30, 2025. Selected users must complete Copilot training. "
        "Policy governs acceptable use and data handling for AI tools. "
        "| Source: https://it.nc.gov/"
    ),
    "Georgia": (
        "YES — 'Red Light, Green Light' guidelines (Aug 2025). "
        "GTA-vetted AI tools: ChatGPT, Google Gemini, Anthropic Claude, Microsoft Copilot approved. "
        "Employees must log in with state credentials. Horizons Innovation Lab / GenAI Sandbox for agency testing. "
        "| Source: https://gta.georgia.gov/"
    ),
    "Colorado": (
        "YES — Free ChatGPT explicitly prohibited on state devices (security policy). "
        "Google Gemini Advanced approved for 150-person pilot → expanded to ~2,000 of 31,000 eligible. "
        "Implies policy framework governing which AI tools are permissible and how data is handled. "
        "| Source: https://oit.colorado.gov/"
    ),
    # Baltimore is a city (will be mapped below)

    # === STATES — researched ===
    "Washington": (
        "PARTIAL — WaTech EA-01-02-S policy governs responsible GenAI use. EO 24-01 (Jan 2024) directs GenAI guidelines. "
        "WaTech GenAI Report (Oct 2024) addresses data governance. No explicit statement that 'AI inputs/outputs are public records' "
        "but WA PRA's broad definition of 'public record' (any writing containing information relating to conduct of government, "
        "RCW 42.56.010) almost certainly covers AI chat logs. "
        "| Source: https://watech.wa.gov/artificial-intelligence-resources"
    ),
    "Massachusetts": (
        "PARTIAL — Enterprise ChatGPT deployment (Feb 2026) for 40K employees includes data governance provisions. "
        "Carahsoft/OpenAI contract includes data retention terms. No standalone AI records retention policy identified. "
        "MA Public Records Law (MGL Ch. 66 § 10) broadly defines records. "
        "| Source: https://www.mass.gov/topics/public-records-requests"
    ),
    "California - CDT (Poppy)": (
        "PARTIAL — CDT GenAI guidelines govern Poppy platform use. Data stays on state network. "
        "Executive Order N-12-23 (Sept 2023) directed agencies to study GenAI risks/benefits. "
        "AB 2885 (2024) requires state agencies to disclose GenAI use in public-facing communications. "
        "No explicit 'AI logs are public records' policy but CPRA broadly covers electronic records. "
        "| Source: https://cdt.ca.gov/technology-innovation/generative-ai/"
    ),
    "Minnesota": (
        "PARTIAL — MNIT GenAI guidelines govern Copilot Chat rollout (April 2025). "
        "nebulaONE chatbot platform has data governance provisions. "
        "MN Data Practices Act (Ch. 13) broadly defines 'government data' to include electronic records. "
        "No standalone AI retention policy identified. "
        "| Source: https://mn.gov/mnit/"
    ),
    "New York": (
        "PARTIAL — ITS AI Pro powered by Google Gemini. Chief AI Officer appointed. "
        "Statewide AI governance framework in development. "
        "NY FOIL broadly defines 'record' (any information kept in any physical form, Public Officers Law § 86). "
        "No standalone AI records retention policy identified. "
        "| Source: https://its.ny.gov/"
    ),
    "Pennsylvania": (
        "UNKNOWN — ChatGPT Enterprise pilot across 14 agencies. "
        "BAD PRECEDENT: OOR ruled ChatGPT logs may be exempt from RTKL. "
        "No AI-specific records retention policy identified. "
        "| Source: OOR ruling (see Prior FOIA Precedent column)"
    ),
    "New Jersey": (
        "PARTIAL — NJ AI Assistant is custom-built; ~20K users, 1M+ prompts. "
        "Platform presumably retains logs by design. "
        "OPRA broadly defines 'government record.' No standalone AI retention policy identified. "
        "| Source: https://innovation.nj.gov/"
    ),
    "Oregon": (
        "PARTIAL — EIS approved Copilot for state employees (Sept 2025). "
        "19-agency AI/GenAI Product & Use Case Inventory published (Apr 2025). "
        "Oregon Public Records Law broadly defines records. No standalone AI retention policy identified. "
        "| Source: https://www.oregon.gov/eis/"
    ),
    "Texas": (
        "PARTIAL — DIR AI guidance exists. TxDOT Copilot deployed to 940+ staff. "
        "30+ active AI initiatives. Texas PIA broadly defines 'public information.' "
        "No standalone AI records retention policy identified. "
        "| Source: https://dir.texas.gov/"
    ),
    "Florida": (
        "UNKNOWN — Limited documented AI deployment. AHCA internal AI for Medicaid analysis. "
        "Florida Sunshine Law is one of the strongest. Constitutional right to records (Art. I, § 24). "
        "No AI-specific records retention policy identified. "
        "| Source: https://dos.fl.gov/sunbiz/"
    ),
    "Utah": (
        "PARTIAL — Google Gemini for Workspace enterprise-wide (15,000-16,000 employees). "
        "DTS likely has AI use guidelines given scale of deployment. "
        "GRAMA broadly defines 'record.' No standalone AI retention policy identified. "
        "| Source: https://dts.utah.gov/"
    ),
    "Ohio": (
        "YES — DAS Policy IT-17 governs statewide AI use. "
        "7 GenAI use cases for disability services in production. Ohio AI Toolkit (Feb 2024). "
        "Policy IT-17 likely includes data governance provisions. "
        "| Source: https://das.ohio.gov/"
    ),
    "Michigan": (
        "PARTIAL — DTMB AI standards published. 'The Bridge' internal AI SharePoint for all employees. "
        "AWS AI Code Generator contract. Michigan FOIA broadly defines 'public record.' "
        "No standalone AI retention policy identified. "
        "| Source: https://www.michigan.gov/dtmb"
    ),
    "Wisconsin": (
        "UNKNOWN — DSPS Maverick AI (Google Cloud + MTX) for licensing. "
        "DWD AI for unemployment claims. No statewide AI policy identified. "
        "WI Open Records Law broadly defines 'record.' "
        "| Source: https://doa.wi.gov/"
    ),
    "New Hampshire": (
        "PARTIAL — DoIT AI policy published. DOJ uses Lexis+ AI; Veterans Home uses ChatGPT. "
        "Small state with dedicated AI officer. "
        "NH Right-to-Know Law broadly defines 'governmental records.' "
        "| Source: https://www.nh.gov/doit/"
    ),
    "Missouri": (
        "PARTIAL — EO 26-02 (Jan 2026) directs AI framework. Google Gemini in secured tenant. "
        "Human-in-the-loop policy for all AI. "
        "Sunshine Law broadly defines 'public record.' No standalone AI retention policy identified. "
        "| Source: https://oa.mo.gov/"
    ),
    "Virginia": (
        "PARTIAL — EO 30 (2024) establishes AI standards. VITA AI policy standards (June 2024). "
        "Copilot Chat available to employees. Agentic AI for regulatory review. "
        "VFOIA broadly defines 'public records.' No explicit 'AI logs are records' statement. "
        "| Source: https://www.vita.virginia.gov/"
    ),
    "Indiana": (
        "PARTIAL — AI Policy v1.1 (Dec 2024) via OCDO addresses data governance for AI. "
        "IARA guidance: AI meeting note outputs must be retained under existing schedules (IC 5-14-1.5-2.9, effective July 2025). "
        "Broad 'writing' definition under APRA likely covers AI chat logs. "
        "| Source: https://www.in.gov/mph/AI/ | https://www.in.gov/iara/"
    ),
    "Montana": (
        "PARTIAL — HB 178 (signed May 2025) requires govt entities to disclose AI use in public interfaces and materials. "
        "Constitutional Art. II, Sec. 9 'Right to Know' broadly covers 'documents' of all public bodies. "
        "No explicit AI retention rule but strong constitutional basis. "
        "| Source: https://leg.mt.gov/bills/2025/billpdf/HB0178.pdf"
    ),
    "Nebraska": (
        "PARTIAL — NITC AI policy section 8-609 (Nov 2024) establishes framework. "
        "Public Records Statutes broadly define records. First 8 hrs staff time free. "
        "No explicit AI retention policy identified. "
        "| Source: https://nitc.nebraska.gov/docs/8-609.pdf"
    ),
    "Maryland": (
        "YES — Governor's EO 01.01.2024.02 (Jan 2024) + AI Governance Act SB818 (July 2024). "
        "AI Subcabinet coordinates policy. Google Gemini extended to 43K employees. "
        "Dedicated AI portal (ai.maryland.gov). AI Enablement Strategy (2025) covers data governance. "
        "MPIA broadly defines 'public record.' "
        "| Source: https://ai.maryland.gov/ | https://doit.maryland.gov/policies/ai/"
    ),
    "Tennessee": (
        "YES — Enterprise AI Policy 200-POL-007 and Generative AI Policy ISC 3.00 published by STS. "
        "AI Review Committee reviews all proposed AI solutions. "
        "AI Advisory Council 2025 Action Plan covers governance. "
        "| Source: https://www.tn.gov/finance/strategic-technology-solutions/artificial-intelligence.html"
    ),

    # === CITIES ===
    "San Francisco, CA": (
        "YES — AI Transparency Ordinance (Ch. 22I, effective 2024). "
        "Requires city departments to disclose AI systems used in decision-making and maintain an AI inventory. "
        "Microsoft Copilot deployed to all ~30K employees (July 2025). "
        "Sunshine Ordinance broadly covers electronic records. "
        "| Source: https://sfgov.org/ai"
    ),
    "San Jose, CA": (
        "PARTIAL — AI & Algorithm Register with vendor FactSheets. "
        "89 ChatGPT Teams licenses for city staff. GovAI Coalition founder. "
        "No standalone AI retention policy identified but register implies documentation. "
        "| Source: https://www.sanjoseca.gov/your-government/departments-offices/information-technology/ai"
    ),
    "Oakland, CA": (
        "PARTIAL — AI guidelines issued Dec 2024 after ChatGPT logs showed PII entered in violation of policy. "
        "~50 employees piloting M365 Copilot. Confirmed: records exist and have been released to journalists. "
        "| Source: https://oaklandside.org (Oct 2025 reporting)"
    ),
    "Bellingham, WA": (
        "UNKNOWN — ChatGPT use confirmed via KNKX investigation. Records obtained by Sanford. "
        "No specific AI retention policy identified beyond general WA PRA. "
        "| Source: https://www.knkx.org/"
    ),
    "Everett, WA": (
        "PARTIAL — Mandated Copilot over ChatGPT for security reasons. "
        "ChatGPT logs obtained via KNKX investigation. "
        "| Source: https://www.knkx.org/"
    ),
    "Kent, WA": (
        "UNKNOWN — Records obtained by Rose Terse from Kent PD in ~30 days. "
        "No specific AI retention policy identified. General WA PRA applies. "
        "| Source: MuckRock"
    ),
    "Bellevue, WA": (
        "UNKNOWN — No specific AI retention policy identified. General WA PRA applies."
    ),
    "Spokane, WA": (
        "UNKNOWN — No specific AI retention policy identified. General WA PRA applies."
    ),
    "Tacoma, WA": (
        "YES — IT Dept published Generative AI Guidelines governing employee use of ChatGPT, Gemini, Copilot. "
        "Guidelines apply to all city employees and third-party contractors. "
        "Existence of policy confirms active use and retention expectations. "
        "| Source: https://www.tacoma.gov/"
    ),
    "New York City": (
        "PARTIAL — OTI published GenAI use guidance (May 2024) allowing agencies to approve tools on agency-by-agency basis. "
        "AI Knowledge Hub launched. Local Law 144 (2021) regulates AI in hiring. "
        "FOIL broadly defines 'record.' "
        "| Source: https://www.nyc.gov/site/oti/"
    ),
    "Washington, DC": (
        "PARTIAL — Microsoft Copilot Chat under DC M365 environment. "
        "Procurement AI (Morfius) for contract management. "
        "DC FOIA broadly defines 'public record.' "
        "| Source: https://octo.dc.gov/"
    ),
    "Kansas City, MO": (
        "UNKNOWN — No confirmed employee GenAI tools. AI efforts are operational (311 routing, pothole detection). "
        "No AI retention policy identified."
    ),
    "Boston, MA": (
        "UNKNOWN — No confirmed enterprise GenAI deployment for city employees. "
        "No AI retention policy identified."
    ),
    "San Diego, CA": (
        "UNKNOWN — No specific AI retention policy identified."
    ),
    "Long Beach, CA": (
        "UNKNOWN — No specific AI retention policy identified."
    ),
    "St. Paul, MN": (
        "UNKNOWN — No specific AI retention policy identified."
    ),
    "Portland, OR (TriMet)": (
        "UNKNOWN — No specific AI retention policy identified."
    ),
    "Houston, TX": (
        "UNKNOWN — No specific AI retention policy identified."
    ),
    "Dallas, TX": (
        "UNKNOWN — No specific AI retention policy identified."
    ),
    "Austin, TX": (
        "UNKNOWN — No specific AI retention policy identified."
    ),
    "Chicago, IL": (
        "PARTIAL — GenAI Roadmap published recommending vetted platforms and center of excellence. "
        "'Shadow AI' reported (workers using free platforms). No formal retention policy. "
        "| Source: https://www.chicago.gov/"
    ),
    "Denver, CO": (
        "PARTIAL — Copilot Chat for city staff. Ask Dee (AiseraGPT) for IT self-service. "
        "CIO title expanded to include AI (2025). "
        "| Source: https://statescoop.com/denver-cio-expanded-role-ai-deployment-caio-2025/"
    ),
    "Atlanta, GA": (
        "PARTIAL — AI Commission (13 members, May 2025). Police body-camera AI confirmed. "
        "No standalone AI retention policy identified. "
        "| Source: https://www.atlantaga.gov/"
    ),
    "Minneapolis, MN": (
        "UNKNOWN — No specific AI retention policy identified."
    ),
    "Tempe, AZ": (
        "UNKNOWN — No specific AI retention policy identified."
    ),
    "Tucson, AZ": (
        "UNKNOWN — No specific AI retention policy identified."
    ),
    # Counties
    "Alameda County, CA": "UNKNOWN — No specific AI retention policy identified.",
    "San Mateo County, CA": "UNKNOWN — No specific AI retention policy identified.",
    "Placer County, CA": "UNKNOWN — No specific AI retention policy identified.",
    "Cook County, IL": "UNKNOWN — No specific AI retention policy identified.",
    "Miami-Dade County, FL": "UNKNOWN — No specific AI retention policy identified.",
    "Broward County, FL": "UNKNOWN — No specific AI retention policy identified.",
    "Clark County, NV": "UNKNOWN — No specific AI retention policy identified.",

    # === FEDERAL ===
    "GSA (USAi.gov)": (
        "PARTIAL — OMB Memo M-24-10 (March 2024) requires agencies to establish AI governance. "
        "NARA Bulletin 2024-02 clarifies AI-generated content is subject to federal records requirements. "
        "USAi.gov platform presumably retains logs. GSA AI strategy includes data governance. "
        "| Source: https://www.whitehouse.gov/omb/briefing-room/2024/03/28/omb-releases-government-wide-policy-to-advance-governance-innovation-and-risk-management-for-agency-use-of-artificial-intelligence/"
    ),
    "HHS": (
        "PARTIAL — OMB M-24-10 applies. ChatGPT Enterprise deployed dept-wide. "
        "Federal Records Act requires retention of records documenting agency activities. "
        "| Source: OMB M-24-10"
    ),
    "OPM": "PARTIAL — OMB M-24-10 applies. Federal Records Act. Copilot + ChatGPT deployed agency-wide.",
    "House of Representatives": "N/A — Congress exempt from FOIA and federal records requirements.",
    "Senate": "N/A — Congress exempt from FOIA and federal records requirements.",
    "DHS (Department of Homeland Security)": (
        "PARTIAL — DHSChat custom platform exists. OMB M-24-10 applies. "
        "CJ Ciaramella filed FOIA for Public Affairs AI logs. "
        "| Source: OMB M-24-10"
    ),
    "SEC (Securities and Exchange Commission)": "PARTIAL — OMB M-24-10 applies. Federal Records Act. Sungho Park filed FOIA.",
    "DOJ (Department of Justice)": "PARTIAL — OMB M-24-10 applies. Dillon Bergin (MuckRock) filed for AI-in-FOIA docs.",
    "EPA (Environmental Protection Agency)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",
    "FDA (Food and Drug Administration)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",
    "NIH (National Institutes of Health)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",
    "CDC (Centers for Disease Control and Prevention)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",
    "USDA (Department of Agriculture)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",
    "DOE (Department of Energy)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",
    "NIST (National Institute of Standards and Technology)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",
    "AMS (Agricultural Marketing Service)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",
    "CMS (Centers for Medicare & Medicaid Services)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",
    "FCC (Federal Communications Commission)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",
    "FEMA (Federal Emergency Management Agency)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",
    "FWS (Fish and Wildlife Service)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",
    "NHTSA (National Highway Traffic Safety Administration)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",
    "NSF (National Science Foundation)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",
    "USCBP (US Customs and Border Protection)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",
    "APHIS (Animal and Plant Health Inspection Service)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",
    "FS (US Forest Service)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",
    "ICE (Immigration and Customs Enforcement)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",
    "NTIA (National Telecommunications and Information Administration)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",
    "USCIS (US Citizenship and Immigration Services)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",
    "FAA (Federal Aviation Administration)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",
    "FDIC (Federal Deposit Insurance Corporation)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",
    "FSIS (Food Safety and Inspection Service)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",
    "MSHA (Mine Safety and Health Administration)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",
    "OSHA (Occupational Safety and Health Administration)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",
    "TSA (Transportation Security Administration)": "PARTIAL — OMB M-24-10 applies. Federal Records Act.",

    # === INTERNATIONAL ===
    "Sweden (10-15 agencies)": (
        "PARTIAL — Sweden's principle of public access to official documents (Offentlighetsprincipen) is constitutional. "
        "AI-generated documents used in government decision-making would be covered. "
        "DIGG (Agency for Digital Government) has AI guidance. "
        "| Source: https://www.digg.se/"
    ),
    "United Kingdom (5-8 depts)": (
        "PARTIAL — Cabinet Office / CDDO published 'Generative AI Framework for HMG' (Jan 2024). "
        "Guidance says AI outputs used in decision-making should be recorded. "
        "FOI Act 2000 broadly covers recorded information. "
        "| Source: https://www.gov.uk/government/publications/generative-ai-framework-for-hmg"
    ),
    "Canada (3-5 depts)": (
        "YES — Treasury Board Directive on Automated Decision-Making (2019, updated 2023). "
        "Requires documentation and records of AI use in government decisions. "
        "ATIA broadly covers records. "
        "| Source: https://www.tbs-sct.canada.ca/pol/doc-eng.aspx?id=32592"
    ),
    "Finland (3-5 agencies)": (
        "PARTIAL — Act on the Openness of Government Activities (621/1999) broadly covers official documents. "
        "AuroraAI program has data governance provisions. "
        "| Source: https://www.finlex.fi/"
    ),
    "Norway (2-3 agencies)": (
        "PARTIAL — Freedom of Information Act (Offentleglova) broadly covers documents. "
        "National AI strategy has governance provisions. "
        "| Source: https://lovdata.no/"
    ),
    "Denmark (2-3 agencies)": (
        "PARTIAL — Access to Public Administration Files Act broadly covers documents. "
        "National AI strategy has governance provisions. "
        "| Source: https://en.digst.dk/"
    ),
    "Estonia": (
        "PARTIAL — Public Information Act broadly covers digital records. "
        "Estonia is a leader in e-governance and digital records. "
        "Kratt (AI strategy) includes governance provisions. "
        "| Source: https://e-estonia.com/"
    ),
    "Netherlands": (
        "PARTIAL — Government Information (Public Access) Act (WOO, 2022) broadly covers documents. "
        "Algorithm Register published. AI governance framework in development. "
        "| Source: https://algoritmeregister.overheid.nl/"
    ),
    "Germany (federal)": (
        "PARTIAL — Federal Freedom of Information Act (IFG) covers official documents. "
        "Federal AI strategy includes governance provisions. "
        "| Source: https://www.bfdi.bund.de/"
    ),
    "France": (
        "PARTIAL — CADA (Commission d'accès aux documents administratifs) governs access. "
        "Code des relations entre le public et l'administration covers administrative documents. "
        "| Source: https://www.cada.fr/"
    ),
    "Australia (federal)": (
        "PARTIAL — FOI Act 1982 covers documents. "
        "Digital Transformation Agency AI guidelines exist. "
        "| Source: https://www.oaic.gov.au/"
    ),
    "New Zealand": (
        "PARTIAL — Official Information Act 1982 covers information held by government. "
        "Algorithm Charter for Aotearoa NZ (2020) — agencies commit to transparency about algorithms. "
        "| Source: https://data.govt.nz/toolkit/data-ethics/government-algorithm-transparency-and-accountability/algorithm-charter/"
    ),
    "EU Institutions": (
        "PARTIAL — Regulation 1049/2001 on public access to documents. "
        "EU AI Act (2024) includes transparency requirements. "
        "| Source: https://www.europarl.europa.eu/"
    ),
    "Mexico": (
        "PARTIAL — General Law of Transparency and Access to Public Information (2015). "
        "INAI oversees access. Digital government strategy includes AI provisions. "
        "| Source: https://www.inai.org.mx/"
    ),
    "South Korea": (
        "PARTIAL — Official Information Disclosure Act broadly covers official documents. "
        "National AI strategy includes governance. "
        "| Source: https://www.mois.go.kr/"
    ),
}

# Apply retention data
for jur_name, data in retention_data.items():
    update_cell(ws, jur_map, jur_name, NEW_COL_RETENTION, data)

print("\nApplied retention policy data")

# =============================================================================
# STEP 5: Fill in Prior FOIA Precedent (Col 24) for ALL rows
# =============================================================================

precedent_data = {
    # === KNOWN PRECEDENTS ===
    "Seattle, WA": (
        "YES — (1) Todd Feathers filed w/ Seattle PD (Aug 2024): slow installments, still processing as of 2025. "
        "Payment required for installments. (2) Rose Terse filed w/ Seattle City Council for specific CM: "
        "'no responsive records.' "
        "| Sources: MuckRock (Todd Feathers, Aug 2024); MuckRock (Rose Terse)"
    ),
    "Bellingham, WA": (
        "YES — COMPLETED. Sanford obtained thousands of pages of ChatGPT logs (~5 months). "
        "KNKX/Cascade PBS investigation (Aug 2025). Records showed staff using ChatGPT for drafting docs, "
        "social media, press releases, image generation. Contract manipulation scandal discovered. "
        "Key insight: calling records officer directly was critical to success. "
        "Records came back as HTML files with Unix timestamps. "
        "| Sources: https://www.knkx.org/government/2025-08-27/washington-state-everett-bellingham-government-officials-embrace-artificial-intelligence-chatgpt-policies-catching-up"
    ),
    "Everett, WA": (
        "YES — COMPLETED. Sanford obtained thousands of pages of ChatGPT logs (~5 months). "
        "KNKX/Cascade PBS investigation (Aug 2025). Dozens of ChatGPT logs obtained. "
        "Use: government documents, social media, policy docs, talking points, image generation. "
        "Records came back as HTML files with Unix timestamps. "
        "| Sources: https://www.knkx.org/"
    ),
    "Kent, WA": (
        "YES — COMPLETED (~30 days). Rose Terse (MuckRock) obtained AI chat records from Kent PD. "
        "Fastest known completion of an AI chat log request. "
        "| Source: MuckRock (Rose Terse)"
    ),
    "Oakland, CA": (
        "YES — Records released to journalists. Oaklandside (Oct 17, 2025) reported: "
        "ChatGPT logs showed staff using ChatGPT for emails; PII entered in violation of policy. "
        "| Source: https://oaklandside.org/"
    ),
    "Pennsylvania": (
        "BAD PRECEDENT — OOR (Office of Open Records) ruled ChatGPT logs may be exempt from RTKL. "
        "This is the most significant adverse ruling for AI chat log requests nationally. "
        "File anyway but be prepared to appeal or distinguish your request. "
        "| Source: PA OOR ruling"
    ),
    "SEC (Securities and Exchange Commission)": (
        "YES — Sungho Park filed FOIA for AI chat logs. Stalled on scope clarification. "
        "| Source: MuckRock"
    ),
    "DHS (Department of Homeland Security)": (
        "YES — CJ Ciaramella filed FOIA for Public Affairs division AI logs. Pending. "
        "| Source: MuckRock"
    ),
    "HHS": (
        "YES — Dillon Bergin (MuckRock) filed for AI-in-FOIA program docs. Status unclear. "
        "| Source: MuckRock"
    ),
    "DOJ (Department of Justice)": (
        "YES — Dillon Bergin (MuckRock) filed for AI-in-FOIA program docs. Status unclear. "
        "| Source: MuckRock"
    ),
    "GSA (USAi.gov)": (
        "UNKNOWN — No known FOIA filed specifically for USAi.gov chat logs. "
        "Auto-deletion of logs reported — FILE IMMEDIATELY."
    ),
    # Baltimore - need to find the right key after row shifts
    "Maryland": (
        "PARTIAL — Todd Feathers reportedly targeting Maryland for AI-related FOIA. "
        "No completed request known at state level. "
        "| Source: MuckRock"
    ),

    # === FEDERAL AGENCIES (MuckRock batch) ===
    # CPSC produced 49 pages
}

# Handle Baltimore which is a specific jurisdiction name issue
# After row insertions, need to find Baltimore row

# Add federal "no known precedent" defaults
federal_no_precedent = [
    "OPM", "AMS (Agricultural Marketing Service)", "CMS (Centers for Medicare & Medicaid Services)",
    "FCC (Federal Communications Commission)", "FEMA (Federal Emergency Management Agency)",
    "FWS (Fish and Wildlife Service)", "NHTSA (National Highway Traffic Safety Administration)",
    "NSF (National Science Foundation)", "USCBP (US Customs and Border Protection)",
    "APHIS (Animal and Plant Health Inspection Service)", "FS (US Forest Service)",
    "ICE (Immigration and Customs Enforcement)", "NIH (National Institutes of Health)",
    "NTIA (National Telecommunications and Information Administration)",
    "USCIS (US Citizenship and Immigration Services)", "CDC (Centers for Disease Control and Prevention)",
    "DOE (Department of Energy)", "FAA (Federal Aviation Administration)",
    "FDIC (Federal Deposit Insurance Corporation)", "FSIS (Food Safety and Inspection Service)",
    "MSHA (Mine Safety and Health Administration)", "NIST (National Institute of Standards and Technology)",
    "OSHA (Occupational Safety and Health Administration)", "TSA (Transportation Security Administration)",
    "USDA (Department of Agriculture)",
    "EPA (Environmental Protection Agency)", "FDA (Food and Drug Administration)",
]
for agency in federal_no_precedent:
    precedent_data[agency] = "UNKNOWN — No known FOIA for AI chat logs at this agency."

# States with no known precedent
states_no_precedent = [
    "Washington", "Massachusetts", "California - CDT (Poppy)", "Minnesota",
    "New York", "Connecticut", "New Jersey", "Colorado", "Oregon", "Texas",
    "Illinois", "Florida", "Utah", "Georgia", "North Carolina", "Ohio",
    "Michigan", "Wisconsin", "New Hampshire",
    "Missouri", "Virginia", "Indiana", "Montana", "Nebraska", "Tennessee",
]
for state in states_no_precedent:
    precedent_data[state] = "UNKNOWN — No known public records request for AI chat logs at state level."

# Cities with no known precedent
cities_no_precedent = [
    "San Francisco, CA", "San Jose, CA", "New York City", "Washington, DC",
    "Kansas City, MO", "Boston, MA", "San Diego, CA", "Long Beach, CA",
    "St. Paul, MN", "Bellevue, WA", "Spokane, WA", "Tacoma, WA",
    "Portland, OR (TriMet)", "Houston, TX", "Dallas, TX", "Austin, TX",
    "Chicago, IL", "Denver, CO", "Atlanta, GA", "Minneapolis, MN",
    "Tempe, AZ", "Tucson, AZ",
]
for city in cities_no_precedent:
    precedent_data[city] = "UNKNOWN — No known public records request for AI chat logs."

# Counties
counties = [
    "Alameda County, CA", "San Mateo County, CA", "Placer County, CA",
    "Cook County, IL", "Miami-Dade County, FL", "Broward County, FL", "Clark County, NV",
]
for county in counties:
    precedent_data[county] = "UNKNOWN — No known public records request for AI chat logs."

# International
intl_no_precedent = [
    "Sweden (10-15 agencies)", "United Kingdom (5-8 depts)", "Canada (3-5 depts)",
    "Finland (3-5 agencies)", "Norway (2-3 agencies)", "Denmark (2-3 agencies)",
    "Estonia", "Netherlands", "Germany (federal)", "France",
    "Australia (federal)", "New Zealand", "EU Institutions", "Mexico", "South Korea",
]
for intl in intl_no_precedent:
    precedent_data[intl] = "UNKNOWN — No known FOI request for AI chat logs."

# Congress
precedent_data["House of Representatives"] = "N/A — Congress exempt from FOIA."
precedent_data["Senate"] = "N/A — Congress exempt from FOIA."

# Apply precedent data
applied_precedent = 0
for jur_name, data in precedent_data.items():
    if update_cell(ws, jur_map, jur_name, NEW_COL_PRECEDENT, data):
        applied_precedent += 1

print(f"Applied precedent data to {applied_precedent} rows")

# =============================================================================
# STEP 6: Fill in Template Adaptation Notes (Col 25) for ALL rows
# =============================================================================

template_notes = {
    # Seattle - cite POL-209
    "Seattle, WA": (
        "Cite POL-209 §6.1-6.3 (AI records are public records, vendors must support export, "
        "employees must retain inputs/outputs). Strongest city-level legal lever. "
        "Specify 'electronic, searchable format' — prior requests involved payment for installments."
    ),
    "Bellingham, WA": (
        "Cite general WA PRA (RCW 42.56). Records came back as HTML files with Unix timestamps — "
        "specify 'electronic, machine-readable format.' Call records officer (Tammy Dixon) directly. "
        "Proven: thousands of pages obtained in ~5 months."
    ),
    "Everett, WA": (
        "Cite general WA PRA. Note: Everett mandated Copilot over ChatGPT for security. "
        "Records came back as HTML files with Unix timestamps. Call records officer directly. "
        "Proven: thousands of pages obtained in ~5 months."
    ),
    "Kent, WA": (
        "Cite general WA PRA. Fastest known completion (~30 days). "
        "Target Kent PD specifically (proven responsive)."
    ),
    "Washington": (
        "Cite WA PRA (RCW 42.56) + EO 24-01 + WaTech EA-01-02-S policy. "
        "File to WaTech for state-level platform logs. Also file to individual agencies. "
        "WA PRA is one of strongest — no exemption for 'burden.'"
    ),
    "Bellevue, WA": "Cite WA PRA (RCW 42.56). General WA template.",
    "Spokane, WA": "Cite WA PRA (RCW 42.56). General WA template.",
    "Tacoma, WA": (
        "Cite WA PRA + Tacoma IT Dept Generative AI Guidelines. "
        "Guidelines confirm ChatGPT, Gemini, Copilot use by employees."
    ),
    "Oakland, CA": (
        "Cite CA PRA. Records already released to Oaklandside journalists. "
        "ChatGPT logs showed PII entered in violation of policy — strong public interest argument."
    ),
    "San Francisco, CA": (
        "Cite CA PRA + AI Transparency Ordinance (Ch. 22I). "
        "Copilot deployed to all ~30K employees. File via NextRequest to Dept of Technology."
    ),
    "San Jose, CA": "Cite CA PRA. Note GovAI Coalition membership. CIO Khaled Tawfik is GovAI board chair.",
    "Connecticut": (
        "Cite CT FOI Act + Policy AI-01 (AI Responsible Use Framework, Feb 2024). "
        "Independent FOI Commission can order disclosure."
    ),
    "Illinois": (
        "Cite IL FOIA + DoIT AI Policy (effective April 2025). "
        "Policy governs state employee AI use — cite it directly."
    ),
    "North Carolina": (
        "Cite NC Public Records Law + Copilot for M365 policy (updated Feb 2025). "
        "Pilot window through June 30, 2025."
    ),
    "Georgia": (
        "Cite GA Open Records Act + 'Red Light, Green Light' guidelines (Aug 2025). "
        "GTA-vetted tools: ChatGPT, Gemini, Claude, Copilot."
    ),
    "Colorado": (
        "Cite CORA + note free ChatGPT is prohibited (only approved tools). "
        "Gemini Advanced approved for ~2,000 users."
    ),
    "Pennsylvania": (
        "Cite PA RTKL. WARNING: OOR ruled ChatGPT logs may be exempt. "
        "Distinguish request or be prepared to appeal. File anyway."
    ),
    "Missouri": (
        "Cite Sunshine Law Ch. 610 RSMo. 3-day deadline. "
        "Mention Google Gemini in secured tenant per EO 26-02."
    ),
    "Virginia": (
        "Cite VA FOIA § 2.2-3700. Mention EO 30 AI standards and VITA Copilot workshops. "
        "5-day deadline with civil penalties."
    ),
    "Indiana": (
        "Cite APRA IC 5-14-3. No fees for electronic records. "
        "Mention AI Policy v1.1 and specific tools (Captain Record, Ask Indiana)."
    ),
    "Montana": (
        "Cite MT Constitution Art. II, Sec. 9 (Right to Know) + MCA 2-6-1002. "
        "Constitutional right is strongest possible lever. "
        "Target Commissioner of Political Practices (documented ChatGPT use)."
    ),
    "Nebraska": (
        "Cite Neb. Rev. Stat. §§ 84-712. 4-day deadline. "
        "8 hours staff time free. Mention NITC AI policy 8-609."
    ),
    "Maryland": (
        "Cite MPIA + Governor's EO 01.01.2024.02 + AI Governance Act SB818. "
        "12,500 active GenAI users. Dedicated AI portal at ai.maryland.gov."
    ),
    "Tennessee": (
        "Cite TN Public Records Act TCA § 10-7-503 + TN Const. Art. I § 19. "
        "Cite Enterprise AI Policy 200-POL-007. 1,000 ChatGPT Enterprise licenses."
    ),
    "Massachusetts": (
        "Cite MA Public Records Law MGL Ch. 66 § 10. "
        "Enterprise ChatGPT for 40K employees — largest state deployment."
    ),
    "California - CDT (Poppy)": (
        "Cite CA PRA + EO N-12-23 + AB 2885 (AI disclosure). "
        "Poppy has 2,348 users across 58 depts. CA is notoriously slow."
    ),
    "Minnesota": (
        "Cite MN Data Practices Act Ch. 13. "
        "Mention Copilot Chat rollout (April 2025) and nebulaONE platform."
    ),
    "New York": (
        "Cite NY FOIL Public Officers Law § 87. "
        "Mention ITS AI Pro (Gemini) pilot of 1,000 workers."
    ),
    "New Jersey": (
        "Cite OPRA. 7-day deadline is very tight. "
        "NJ AI Assistant has 20K users, 1M+ prompts — specific target."
    ),
    "Oregon": "Cite Oregon Public Records Law. Mention EIS Copilot approval (Sept 2025).",
    "Texas": "Cite TX PIA. Mention TxDOT Copilot (940+ staff) and 30+ AI initiatives.",
    "Florida": "Cite FL Sunshine Law + Art. I § 24 (constitutional right). One of strongest sunshine laws.",
    "Utah": (
        "Cite GRAMA. Gemini for Workspace at 15,000-16,000 employees (~40% adoption). "
        "Centralized GRAMA portal."
    ),
    "Ohio": "Cite OH Public Records Act + DAS Policy IT-17. 7 GenAI use cases in production.",
    "Michigan": "Cite MI FOIA. Mention DTMB AI standards and 'The Bridge' internal AI site.",
    "Wisconsin": "Cite WI Open Records Law. Mention DSPS Maverick AI and DWD AI tools.",
    "New Hampshire": "Cite NH Right-to-Know Law RSA 91-A. Small state with dedicated AI officer.",
    "New York City": "Cite NYC FOIL. File to DoITT/OTI for AI tools. Mention GenAI guidance (May 2024).",
    "Washington, DC": "Cite DC FOIA. File via myfoiadc.govqa.us. Mention Copilot Chat in DC M365.",
    "Kansas City, MO": "Cite MO Sunshine Law. 3-day deadline. Limited AI deployment.",
    "Boston, MA": "Cite MA Public Records Law. Limited confirmed AI deployment.",
}

# Federal template notes
fed_template = "Cite FOIA 5 U.S.C. § 552. Reference OMB M-24-10 (AI governance) and NARA Bulletin 2024-02 (AI records retention)."
for jur_name in jur_map:
    if jur_name not in template_notes:
        row_cat = ws.cell(row=jur_map[jur_name], column=1).value
        if row_cat == "US-Fed":
            template_notes[jur_name] = fed_template
        elif row_cat in ("Intl", "Intl2"):
            template_notes[jur_name] = "Cite applicable FOI statute. Adapt to local language and legal framework."
        elif row_cat == "US-City":
            # Get the state from jurisdiction name
            template_notes[jur_name] = "Cite applicable state public records law."
        elif row_cat == "US-County":
            template_notes[jur_name] = "Cite applicable state public records law."

# Apply template notes
applied_template = 0
for jur_name, data in template_notes.items():
    if update_cell(ws, jur_map, jur_name, NEW_COL_TEMPLATE, data):
        applied_template += 1

print(f"Applied template notes to {applied_template} rows")

# =============================================================================
# STEP 7: Set column widths for new columns
# =============================================================================
ws.column_dimensions['W'].width = 50
ws.column_dimensions['X'].width = 50
ws.column_dimensions['Y'].width = 50

# =============================================================================
# STEP 8: Save
# =============================================================================
save_tracker(wb, "foia_tracker_detailed_v3.xlsx")
print(f"Total rows: {ws.max_row} (was 102, now {ws.max_row})")

# Verify
from helpers import load_tracker as _lt
wb2, ws2 = _lt("foia_tracker_detailed_v3.xlsx")
print(f"\nVerification:")
print(f"  Headers: {[ws2.cell(row=1, column=c).value for c in range(23, 26)]}")
print(f"  Row 26 (should be Washington): {ws2.cell(row=26, column=3).value}")
print(f"  Row 33 (should be Tennessee): {ws2.cell(row=33, column=3).value}")
# Find Kent
for r in range(2, ws2.max_row + 1):
    if ws2.cell(row=r, column=3).value == "Kent, WA":
        print(f"  Kent, WA found at row {r}")
        break
# Spot check new columns
print(f"  Seattle retention policy (first 80 chars): {str(ws2.cell(row=jur_map.get('Seattle, WA', 39), column=23).value)[:80]}")
print(f"  PA precedent (first 80 chars): {str(ws2.cell(row=jur_map.get('Pennsylvania', 19), column=24).value)[:80]}")
