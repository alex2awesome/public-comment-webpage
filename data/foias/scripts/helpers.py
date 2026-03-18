#!/usr/bin/env python3
"""Shared utilities for FOIA tracker enrichment scripts."""

import openpyxl
from openpyxl.styles import Font, Alignment
from pathlib import Path

TRACKER_DIR = Path(__file__).parent.parent / "tracking-spreadsheets"

# Column indices (1-indexed, matching openpyxl)
COL_CAT = 1
COL_TIER = 2
COL_JURISDICTION = 3
COL_COUNTRY = 4
COL_LEVEL = 5
COL_AI_DEPLOY = 6
COL_EMPLOYEES = 7
COL_PROXY = 8
COL_FILING = 9
COL_SOURCE_URL = 10
COL_COST = 11
COL_STATUTE = 12
COL_DEADLINE = 13
COL_TURNAROUND = 14
COL_TEMPLATE = 15
COL_GOVAI = 16
COL_STATUS = 17
COL_DATE_FILED = 18
COL_DATE_DUE = 19
COL_DATE_RECEIVED = 20
COL_NOTES = 21
COL_REQUIREMENTS = 22
COL_RETENTION = 23
COL_PRECEDENT = 24
COL_TEMPLATE_NOTES = 25

COLUMN_KEYS = [
    "cat", "tier", "jurisdiction", "country", "level",
    "ai_deploy", "employees", "proxy", "filing", "source_url",
    "cost", "statute", "deadline", "turnaround", "template",
    "govai", "status", "date_filed", "date_due", "date_received",
    "notes", "requirements", "retention", "precedent", "template_notes",
]


def load_tracker(filename):
    """Load a tracker spreadsheet and return (workbook, worksheet)."""
    path = TRACKER_DIR / filename
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    return wb, ws


def save_tracker(wb, filename):
    """Save a tracker spreadsheet."""
    path = TRACKER_DIR / filename
    wb.save(str(path))
    print(f"Saved to {path}")


def build_jur_map(ws):
    """Build jurisdiction name -> row number mapping."""
    jur_map = {}
    for row in range(2, ws.max_row + 1):
        jur = ws.cell(row=row, column=COL_JURISDICTION).value
        if jur:
            jur_map[jur] = row
    return jur_map


def update_cell(ws, jur_map, jurisdiction, col, value):
    """Update a cell by jurisdiction name and column index."""
    if jurisdiction in jur_map:
        row = jur_map[jurisdiction]
        ws.cell(row=row, column=col, value=value)
        ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True)
        return True
    print(f"  WARNING: '{jurisdiction}' not found in spreadsheet")
    return False


def append_notes(ws, jur_map, jurisdiction, new_text,
                 header="FILING GUIDANCE (BROAD ALL-EMPLOYEE REQUEST)"):
    """Append to existing notes with separator."""
    if jurisdiction not in jur_map:
        print(f"  WARNING: '{jurisdiction}' not found")
        return False
    row = jur_map[jurisdiction]
    existing = ws.cell(row, COL_NOTES).value or ""
    if existing:
        combined = existing.rstrip() + f"\n\n---\n{header}:\n" + new_text
    else:
        combined = f"{header}:\n" + new_text
    ws.cell(row, COL_NOTES, combined)
    return True


def apply_updates(ws, jur_map, updates_dict, col):
    """Apply a dict of {jurisdiction: value} updates to a column. Returns count."""
    count = 0
    for jur, val in updates_dict.items():
        if update_cell(ws, jur_map, jur, col, val):
            count += 1
    return count


def add_agency_row(ws, row_num, agency_dict):
    """Write a new agency row from a dict with standard column keys."""
    for i, key in enumerate(COLUMN_KEYS):
        ws.cell(row_num, i + 1, agency_dict.get(key, ""))
